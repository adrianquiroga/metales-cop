import os
import sys
import json
import datetime
import urllib.request
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.worksheet.datavalidation import DataValidation

# Constante de conversión: 1 Onza Troy = 31.1034768 gramos
TROY_OUNCE_TO_GRAMS = 31.1034768

# Fecha de inicio requerida: 1 de Enero de 2025
DEFAULT_START_DATE = datetime.date(2025, 1, 1)

# Rutas de archivos
script_dir = os.path.dirname(os.path.abspath(__file__))
excel_path = os.path.join(script_dir, "Precios_Metales_COP.xlsx")
data_js_path = os.path.join(script_dir, "data.js")
data_json_path = os.path.join(script_dir, "data.json")

headers_http = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
}

def fetch_yahoo_history(ticker, start_date, end_date):
    """
    Obtiene el historial diario de cierre para un ticker de Yahoo Finance entre start_date y end_date.
    """
    p1 = int(datetime.datetime.combine(start_date, datetime.time.min, tzinfo=datetime.timezone.utc).timestamp())
    p2 = int(datetime.datetime.combine(end_date + datetime.timedelta(days=1), datetime.time.max, tzinfo=datetime.timezone.utc).timestamp())
    
    url = f"https://query1.finance.yahoo.com/v8/finance/chart/{ticker}?period1={p1}&period2={p2}&interval=1d"
    req = urllib.request.Request(url, headers=headers_http)
    
    try:
        with urllib.request.urlopen(req, timeout=15) as response:
            data = json.loads(response.read().decode())
            result = data.get('chart', {}).get('result')
            if not result:
                return {}
            
            timestamps = result[0].get('timestamp', [])
            quote = result[0].get('indicators', {}).get('quote', [{}])[0]
            closes = quote.get('close', [])
            
            history = {}
            for ts, close in zip(timestamps, closes):
                if close is not None:
                    dt_str = datetime.datetime.fromtimestamp(ts, tz=datetime.timezone.utc).strftime('%Y-%m-%d')
                    history[dt_str] = float(close)
            return history
    except Exception as e:
        print(f"  [ERROR] No se pudo obtener datos para {ticker}: {e}")
        return {}

def get_excel_dates_info(ws):
    """
    Analiza las fechas registradas en la Columna A (Fecha).
    Retorna (min_date, max_date, existing_dates_dict).
    """
    min_date = None
    max_date = None
    existing_dates = set()
    
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val:
            dt = None
            if isinstance(val, datetime.datetime):
                dt = val.date()
            elif isinstance(val, datetime.date):
                dt = val
            elif isinstance(val, str):
                try:
                    dt = datetime.datetime.strptime(val.strip(), '%Y-%m-%d').date()
                except ValueError:
                    continue
            
            if dt:
                dt_str = dt.strftime('%Y-%m-%d')
                existing_dates.add(dt_str)
                if min_date is None or dt < min_date:
                    min_date = dt
                if max_date is None or dt > max_date:
                    max_date = dt
                    
    return min_date, max_date, existing_dates

def export_json_js_data(ws_data):
    """
    Exporta la serie histórica completa a data.js y data.json para la aplicación web interactiva.
    """
    json_list = []
    max_r = ws_data.max_row
    
    for r in range(2, max_r + 1):
        val_date = ws_data.cell(row=r, column=1).value
        if not val_date:
            continue
            
        if isinstance(val_date, (datetime.date, datetime.datetime)):
            d_str = val_date.strftime('%Y-%m-%d')
        else:
            d_str = str(val_date).strip()
            
        g_usd = ws_data.cell(row=r, column=2).value or 0
        s_usd = ws_data.cell(row=r, column=3).value or 0
        trm = ws_data.cell(row=r, column=4).value or 0
        
        g_cop = round((g_usd / TROY_OUNCE_TO_GRAMS) * trm, 2)
        s_cop = round((s_usd / TROY_OUNCE_TO_GRAMS) * trm, 2)
        
        json_list.append({
            "fecha": d_str,
            "oro_usd": round(g_usd, 2),
            "plata_usd": round(s_usd, 2),
            "trm": round(trm, 2),
            "oro_cop_g": g_cop,
            "plata_cop_g": s_cop
        })
        
    with open(data_json_path, "w", encoding="utf-8") as f:
        json.dump(json_list, f, indent=2)
        
    with open(data_js_path, "w", encoding="utf-8") as f:
        f.write("window.METALES_DATA = " + json.dumps(json_list, indent=2) + ";")
        
    print(f"Exportados {len(json_list)} registros a 'data.js' y 'data.json' para la app web.")

def build_dashboard_sheet(wb, ws_data):
    """
    Crea o recrea la pestaña de Dashboard asegurando visibilidad total del eje Y y sin solapamientos.
    """
    sheet_name = "Dashboard y Gráficos"
    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
        
    ws_dash = wb.create_sheet(sheet_name, index=0)
    ws_dash.views.sheetView[0].showGridLines = True
    
    # Estilos
    font_name = "Segoe UI"
    fill_dark = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    fill_kpi = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color="E2E8F0"),
        right=Side(style='thin', color="E2E8F0"),
        top=Side(style='thin', color="E2E8F0"),
        bottom=Side(style='thin', color="E2E8F0")
    )
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    
    # 1. Título Banner
    ws_dash.merge_cells("A1:N1")
    title_cell = ws_dash["A1"]
    title_cell.value = "MONITOREO DE METALES PRECIOSOS: ORO Y PLATA EN PESOS COLOMBIANOS (COP)"
    title_cell.font = Font(name=font_name, size=13, bold=True, color="FFFFFF")
    title_cell.fill = fill_dark
    title_cell.alignment = align_center
    ws_dash.row_dimensions[1].height = 36

    # 2. Tarjetas KPI (Filas 3 a 5)
    max_r = ws_data.max_row
    
    kpis = {
        "A3": ("ÚLTIMA FECHA", f"='Precios Diarios'!A{max_r}", "YYYY-MM-DD"),
        "D3": ("ORO (COP/g)", f"='Precios Diarios'!E{max_r}", "$ #,##0.00"),
        "G3": ("PLATA (COP/g)", f"='Precios Diarios'!F{max_r}", "$ #,##0.00"),
        "J3": ("TRM (USD/COP)", f"='Precios Diarios'!D{max_r}", "$ #,##0.00"),
        "M3": ("VAR. DÍA ORO", f"='Precios Diarios'!G{max_r}", "0.00%")
    }
    
    for pos, (lbl, formula, num_fmt) in kpis.items():
        col_let = pos[0]
        col_idx = ord(col_let) - 64
        
        if pos == "A3":
            ws_dash.merge_cells("A3:C3")
            ws_dash.merge_cells("A4:C5")
        elif pos == "D3":
            ws_dash.merge_cells("D3:F3")
            ws_dash.merge_cells("D4:F5")
        elif pos == "G3":
            ws_dash.merge_cells("G3:I3")
            ws_dash.merge_cells("G4:I5")
        elif pos == "J3":
            ws_dash.merge_cells("J3:L3")
            ws_dash.merge_cells("J4:L5")
        elif pos == "M3":
            ws_dash.merge_cells("M3:N3")
            ws_dash.merge_cells("M4:N5")
            
        lbl_cell = ws_dash.cell(row=3, column=col_idx)
        lbl_cell.value = lbl
        lbl_cell.font = Font(name=font_name, size=9, bold=True, color="64748B")
        lbl_cell.alignment = align_center
        lbl_cell.fill = fill_kpi
        
        val_cell = ws_dash.cell(row=4, column=col_idx)
        val_cell.value = formula
        val_cell.font = Font(name=font_name, size=15, bold=True, color="1E293B")
        val_cell.number_format = num_fmt
        val_cell.alignment = align_center
        val_cell.fill = fill_kpi

    # Bordes para Tarjetas KPI
    for r in range(3, 6):
        for c in range(1, 15):
            ws_dash.cell(row=r, column=c).border = thin_border

    # 3. Control Selector de Periodo Interactivo (Filas 7 a 9)
    ws_dash.merge_cells("A7:D7")
    sel_lbl = ws_dash["A7"]
    sel_lbl.value = "SELECCIÓN DE PERÍODO A GRAFICAR:"
    sel_lbl.font = Font(name=font_name, size=10, bold=True, color="1E293B")
    sel_lbl.alignment = align_left
    
    ws_dash.merge_cells("A8:D8")
    sel_cell = ws_dash["A8"]
    sel_cell.value = "Todo el Historial (Desde 2025)"
    sel_cell.font = Font(name=font_name, size=11, bold=True, color="0F172A")
    sel_cell.fill = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")
    sel_cell.alignment = align_center
    sel_cell.border = thin_border
    ws_dash.row_dimensions[8].height = 24
    
    # Validación de datos desplegable en A8
    dv = DataValidation(
        type="list", 
        formula1='"Todo el Historial (Desde 2025),Año 2025,Año 2026,Últimos 90 Días,Últimos 30 Días"',
        allow_blank=False
    )
    ws_dash.add_data_validation(dv)
    dv.add(ws_dash["A8"])
    
    # Celdas auxiliares E8, F8, G8, H8, I8
    ws_dash["E8"] = '=IF(A8="Año 2025", DATE(2025,1,1), IF(A8="Año 2026", DATE(2026,1,1), IF(A8="Últimos 90 Días", TODAY()-90, IF(A8="Últimos 30 Días", TODAY()-30, DATE(2025,1,1)))))'
    ws_dash["F8"] = '=IF(A8="Año 2025", DATE(2025,12,31), IF(A8="Año 2026", DATE(2026,12,31), TODAY()))'
    
    ws_dash["G8"] = f'=IFERROR(MATCH(E8, \'Precios Diarios\'!$A$2:$A${max_r}, 1), 1) + 1'
    ws_dash["H8"] = f'=IFERROR(MATCH(F8, \'Precios Diarios\'!$A$2:$A${max_r}, 1), {max_r-1}) + 1'
    ws_dash["I8"] = '=MAX(1, H8 - G8 + 1)'
    
    ws_dash["E8"].number_format = "YYYY-MM-DD"
    ws_dash["F8"].number_format = "YYYY-MM-DD"
    for c_h in ["E8", "F8", "G8", "H8", "I8"]:
        ws_dash[c_h].font = Font(name=font_name, size=8, color="94A3B8")
    
    note_cell = ws_dash.cell(row=9, column=1, value="Nota: Para cambiar el período use A8. Para ver precios exactos al pasar el cursor, abra 'index.html'.")
    note_cell.font = Font(name=font_name, size=8, italic=True, color="64748B")

    # 4. Tabla de Datos Re-indexada Dinámicamente (Columnas P, Q, R)
    ws_dash.cell(row=1, column=16, value="Fecha")
    ws_dash.cell(row=1, column=17, value="Oro (COP/g)")
    ws_dash.cell(row=1, column=18, value="Plata (COP/g)")
    
    for k in range(1, max_r):
        r = k + 1
        cell_f = ws_dash.cell(row=r, column=16, value=f"=IF({k}<=$I$8, INDEX('Precios Diarios'!$A$2:$A${max_r}, $G$8 - 1 + {k} - 1), NA())")
        cell_f.number_format = "YYYY-MM-DD"
        
        ws_dash.cell(row=r, column=17, value=f"=IF({k}<=$I$8, INDEX('Precios Diarios'!$E$2:$E${max_r}, $G$8 - 1 + {k} - 1), NA())")
        ws_dash.cell(row=r, column=18, value=f"=IF({k}<=$I$8, INDEX('Precios Diarios'!$F$2:$F${max_r}, $G$8 - 1 + {k} - 1), NA())")

    # 5. Construcción de Gráficos Excel con eje Y amplio y sin solapamiento
    data_gold = Reference(ws_dash, min_col=17, min_row=1, max_row=max_r)
    data_silver = Reference(ws_dash, min_col=18, min_row=1, max_row=max_r)
    cats = Reference(ws_dash, min_col=16, min_row=2, max_row=max_r)

    # Gráfico 1: Precio del Oro en COP/g
    chart_gold = LineChart()
    chart_gold.title = "Evolución del Precio del ORO (COP / Gramo)"
    chart_gold.style = 13
    chart_gold.width = 17
    chart_gold.height = 11
    chart_gold.legend = None # Sin caja de leyenda lateral para dejar espacio al eje Y

    chart_gold.y_axis.delete = False
    chart_gold.y_axis.title = "COP / g"
    chart_gold.y_axis.tickLblPos = "nextTo"
    chart_gold.y_axis.number_format = '"$"#,##0'
    chart_gold.y_axis.majorGridlines = ChartLines()

    chart_gold.x_axis.delete = False
    chart_gold.x_axis.title = "Fecha"
    chart_gold.x_axis.tickLblPos = "nextTo"

    chart_gold.add_data(data_gold, titles_from_data=True, from_rows=False)
    chart_gold.set_categories(cats)

    if chart_gold.series:
        s1 = chart_gold.series[0]
        s1.graphicalProperties.line.solidFill = "D97706" # Dorado
        s1.graphicalProperties.line.width = 22000
        s1.marker.symbol = "circle"
        s1.marker.size = 3

    ws_dash.add_chart(chart_gold, "A11")

    # Gráfico 2: Precio de la Plata en COP/g
    chart_silver = LineChart()
    chart_silver.title = "Evolución del Precio de la PLATA (COP / Gramo)"
    chart_silver.style = 13
    chart_silver.width = 17
    chart_silver.height = 11
    chart_silver.legend = None

    chart_silver.y_axis.delete = False
    chart_silver.y_axis.title = "COP / g"
    chart_silver.y_axis.tickLblPos = "nextTo"
    chart_silver.y_axis.number_format = '"$"#,##0'
    chart_silver.y_axis.majorGridlines = ChartLines()

    chart_silver.x_axis.delete = False
    chart_silver.x_axis.title = "Fecha"
    chart_silver.x_axis.tickLblPos = "nextTo"

    chart_silver.add_data(data_silver, titles_from_data=True, from_rows=False)
    chart_silver.set_categories(cats)

    if chart_silver.series:
        s2 = chart_silver.series[0]
        s2.graphicalProperties.line.solidFill = "475569" # Plata / Slate
        s2.graphicalProperties.line.width = 22000
        s2.marker.symbol = "circle"
        s2.marker.size = 3

    ws_dash.add_chart(chart_silver, "H11")

    # Gráfico 3: Comparativo Combinado Eje Doble (Oro vs Plata COP/g)
    c_comb_gold = LineChart()
    c_comb_gold.title = "Comparativo Histórico: ORO vs PLATA (COP/g) [Eje Doble]"
    c_comb_gold.style = 13
    c_comb_gold.width = 34
    c_comb_gold.height = 12

    c_comb_gold.y_axis.delete = False
    c_comb_gold.y_axis.title = "Oro (COP/g)"
    c_comb_gold.y_axis.tickLblPos = "nextTo"
    c_comb_gold.y_axis.number_format = '"$"#,##0'
    c_comb_gold.y_axis.majorGridlines = ChartLines()

    c_comb_gold.x_axis.delete = False
    c_comb_gold.x_axis.title = "Fecha"
    c_comb_gold.x_axis.tickLblPos = "nextTo"

    c_comb_gold.add_data(data_gold, titles_from_data=True, from_rows=False)
    c_comb_gold.set_categories(cats)
    if c_comb_gold.series:
        c_comb_gold.series[0].graphicalProperties.line.solidFill = "D97706"

    c_comb_silver = LineChart()
    c_comb_silver.y_axis.delete = False
    c_comb_silver.y_axis.title = "Plata (COP/g)"
    c_comb_silver.y_axis.tickLblPos = "nextTo"
    c_comb_silver.y_axis.number_format = '"$"#,##0'
    c_comb_silver.y_axis.axId = 200
    c_comb_silver.y_axis.crosses = "max"

    c_comb_silver.add_data(data_silver, titles_from_data=True, from_rows=False)
    if c_comb_silver.series:
        c_comb_silver.series[0].graphicalProperties.line.solidFill = "475569"

    c_comb_gold += c_comb_silver
    # Ubicar la leyenda en la parte inferior ("b") para que NUNCA choque con el título del gráfico
    c_comb_gold.legend.position = "b"
    
    ws_dash.add_chart(c_comb_gold, "A27")

    # Ajustar dimensiones de columnas
    for c in range(1, 16):
        ws_dash.column_dimensions[get_column_letter(c)].width = 12

def main():
    print("==================================================")
    print("  ACTUALIZADOR DE METALES (ORO Y PLATA EN COP)    ")
    print("   Historial Completo desde 1 de Enero de 2025    ")
    print("==================================================")
    
    today = datetime.date.today()
    file_exists = os.path.exists(excel_path)
    
    rebuild_required = False
    
    if file_exists:
        print(f"Cargando libro existente: '{os.path.basename(excel_path)}'...")
        wb = openpyxl.load_workbook(excel_path)
        if "Precios Diarios" in wb.sheetnames:
            ws_data = wb["Precios Diarios"]
        else:
            ws_data = wb.active
            
        min_date, max_date, existing_dates = get_excel_dates_info(ws_data)
        
        if min_date is None or min_date > DEFAULT_START_DATE:
            print(f"Detectado historial incompleto (inicio actual: {min_date}). Reconstruyendo desde 2025-01-01...")
            rebuild_required = True
            start_date = DEFAULT_START_DATE
        elif max_date and max_date >= today:
            print("Los datos ya se encuentran actualizados hasta el dia de hoy.")
            build_dashboard_sheet(wb, ws_data)
            export_json_js_data(ws_data)
            wb.save(excel_path)
            print("Dashboard Excel con eje Y visible y App Web actualizados con exito.")
            return
        else:
            start_date = max_date + datetime.timedelta(days=1)
            print(f"Ultima fecha registrada: {max_date}. Actualizando desde {start_date} hasta {today}...")
    else:
        print("Creando nuevo libro de Excel desde el 1 de Enero de 2025...")
        rebuild_required = True
        start_date = DEFAULT_START_DATE

    if rebuild_required:
        wb = openpyxl.Workbook()
        ws_data = wb.active
        ws_data.title = "Precios Diarios"
        ws_data.views.sheetView[0].showGridLines = True
        existing_dates = set()

    # Consultar datos en Yahoo Finance
    print(f"\nConsultando cotizaciones internacionales desde {start_date} hasta {today}...")
    print("  -> Oro (GC=F)...")
    gold_data = fetch_yahoo_history("GC=F", start_date - datetime.timedelta(days=7), today)
    print("  -> Plata (SI=F)...")
    silver_data = fetch_yahoo_history("SI=F", start_date - datetime.timedelta(days=7), today)
    print("  -> TRM (USDCOP=X)...")
    trm_data = fetch_yahoo_history("USDCOP=X", start_date - datetime.timedelta(days=7), today)
    
    # Obtener todas las fechas ordenadas
    all_dates = sorted(set(gold_data.keys()).union(set(silver_data.keys())).union(set(trm_data.keys())))
    filter_dates = [d for d in all_dates if datetime.datetime.strptime(d, '%Y-%m-%d').date() >= start_date]
    
    if not filter_dates and not rebuild_required:
        print("No se encontraron nuevos registros.")
        return

    # Estilos visuales
    font_name = "Segoe UI"
    font_header = Font(name=font_name, size=10, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color="E2E8F0"),
        right=Side(style='thin', color="E2E8F0"),
        top=Side(style='thin', color="E2E8F0"),
        bottom=Side(style='thin', color="E2E8F0")
    )
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    # Encabezados en Precios Diarios
    if rebuild_required or ws_data.max_row == 1:
        headers = [
            "Fecha", 
            "Oro (USD/oz)", 
            "Plata (USD/oz)", 
            "TRM (USD/COP)", 
            "Oro (COP/g)", 
            "Plata (COP/g)", 
            "Var. Oro %", 
            "Var. Plata %"
        ]
        ws_data.append(headers)
        ws_data.row_dimensions[1].height = 26
        for col_num in range(1, len(headers) + 1):
            cell = ws_data.cell(row=1, column=col_num)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = thin_border

    # Precios anteriores para forward-fill
    last_gold = None
    last_silver = None
    last_trm = None
    
    if ws_data.max_row > 1 and not rebuild_required:
        prev_r = ws_data.max_row
        last_gold = ws_data.cell(row=prev_r, column=2).value
        last_silver = ws_data.cell(row=prev_r, column=3).value
        last_trm = ws_data.cell(row=prev_r, column=4).value

    # Insertar filas
    print(f"Insertando {len(filter_dates)} registros en la hoja 'Precios Diarios'...")
    rows_added = 0
    
    for d_str in filter_dates:
        if d_str in existing_dates and not rebuild_required:
            continue
            
        g_val = gold_data.get(d_str, last_gold)
        s_val = silver_data.get(d_str, last_silver)
        t_val = trm_data.get(d_str, last_trm)
        
        if g_val is None or s_val is None or t_val is None:
            continue
            
        last_gold, last_silver, last_trm = g_val, s_val, t_val
        
        curr_row = ws_data.max_row + 1
        ws_data.row_dimensions[curr_row].height = 20
        
        dt_obj = datetime.datetime.strptime(d_str, '%Y-%m-%d').date()
        
        cell_date = ws_data.cell(row=curr_row, column=1, value=dt_obj)
        cell_date.number_format = "YYYY-MM-DD"
        cell_date.alignment = align_center
        cell_date.font = Font(name=font_name, size=10)
        cell_date.border = thin_border
        
        cell_g_usd = ws_data.cell(row=curr_row, column=2, value=g_val)
        cell_g_usd.number_format = "$ #,##0.00"
        cell_g_usd.alignment = align_right
        cell_g_usd.font = Font(name=font_name, size=10)
        cell_g_usd.border = thin_border
        
        cell_s_usd = ws_data.cell(row=curr_row, column=3, value=s_val)
        cell_s_usd.number_format = "$ #,##0.00"
        cell_s_usd.alignment = align_right
        cell_s_usd.font = Font(name=font_name, size=10)
        cell_s_usd.border = thin_border
        
        cell_trm = ws_data.cell(row=curr_row, column=4, value=t_val)
        cell_trm.number_format = "$ #,##0.00"
        cell_trm.alignment = align_right
        cell_trm.font = Font(name=font_name, size=10)
        cell_trm.border = thin_border
        
        cell_g_cop = ws_data.cell(row=curr_row, column=5, value=f"=ROUND((B{curr_row}/{TROY_OUNCE_TO_GRAMS})*D{curr_row}, 2)")
        cell_g_cop.number_format = "$ #,##0.00"
        cell_g_cop.alignment = align_right
        cell_g_cop.font = Font(name=font_name, size=10, bold=True)
        cell_g_cop.border = thin_border
        
        cell_s_cop = ws_data.cell(row=curr_row, column=6, value=f"=ROUND((C{curr_row}/{TROY_OUNCE_TO_GRAMS})*D{curr_row}, 2)")
        cell_s_cop.number_format = "$ #,##0.00"
        cell_s_cop.alignment = align_right
        cell_s_cop.font = Font(name=font_name, size=10, bold=True)
        cell_s_cop.border = thin_border

        if curr_row > 2:
            cell_g_var = ws_data.cell(row=curr_row, column=7, value=f"=IFERROR((E{curr_row}-E{curr_row-1})/E{curr_row-1}, 0)")
            cell_s_var = ws_data.cell(row=curr_row, column=8, value=f"=IFERROR((F{curr_row}-F{curr_row-1})/F{curr_row-1}, 0)")
        else:
            cell_g_var = ws_data.cell(row=curr_row, column=7, value=0)
            cell_s_var = ws_data.cell(row=curr_row, column=8, value=0)
            
        cell_g_var.number_format = "0.00%"
        cell_g_var.alignment = align_right
        cell_g_var.font = Font(name=font_name, size=10)
        cell_g_var.border = thin_border
        
        cell_s_var.number_format = "0.00%"
        cell_s_var.alignment = align_right
        cell_s_var.font = Font(name=font_name, size=10)
        cell_s_var.border = thin_border

        if curr_row % 2 == 1:
            for c_idx in range(1, 9):
                ws_data.cell(row=curr_row, column=c_idx).fill = fill_zebra
                
        rows_added += 1

    # Ajustar anchos de columna en Precios Diarios
    col_widths = {1: 14, 2: 16, 3: 16, 4: 16, 5: 18, 6: 18, 7: 14, 8: 14}
    for col_idx, width in col_widths.items():
        col_letter = get_column_letter(col_idx)
        ws_data.column_dimensions[col_letter].width = width

    # Habilitar AutoFiltro en Precios Diarios
    ws_data.auto_filter.ref = f"A1:H{ws_data.max_row}"

    # Reconstruir Dashboard Excel con eje Y totalmente visible
    print("Generando pestana 'Dashboard y Graficos' con eje Y visible...")
    build_dashboard_sheet(wb, ws_data)
    
    # Exportar datos para la web app
    export_json_js_data(ws_data)

    # Guardar Excel
    print(f"Guardando cambios en '{os.path.basename(excel_path)}'...")
    try:
        wb.save(excel_path)
        print(f"Exito! Se actualizaron correctamente {ws_data.max_row - 1} registros diarios.")
    except Exception as e:
        print(f"[ERROR] No se pudo guardar el archivo Excel: {e}")

if __name__ == "__main__":
    main()
